import streamlit as st
import os
from datetime import datetime, time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import pandas as pd
import numpy as np
from urllib.request import urlopen
from PIL import Image
from io import BytesIO

# ==== HEADER PERSONALIZADO ====
st.set_page_config(page_title="Reporte de Tipo de Cambio", layout="wide")

# --- logo en la izquierda ---
col1, col2 = st.columns([1, 6])
with col1:
    try:
        url_logo = "https://i.pinimg.com/736x/fa/36/b0/fa36b02d7c35643ac5c45c03a7274b20.jpg" 
        image = Image.open(urlopen(url_logo))
        st.image(image, width=180)

    except:
        st.write("")  # fallback si falla

with col2:
    st.markdown("<h1 style='text-align:left; color:#1F4E79;'>Reporte de Tipo de Cambio</h1>", unsafe_allow_html=True)
    st.markdown("<h4 style='text-align:left; color:#4B4B4B;'>Franco Olivares - Strategic Asset Allocation</h4>", unsafe_allow_html=True)

st.markdown("---")

# ==== Funciones (parse_hora_to_time y generar_reporte_streamlit) ====
def parse_hora_to_time(s):
    if pd.isna(s):
        return None
    if isinstance(s, time):
        return s
    if isinstance(s, datetime):
        return s.time()
    s = str(s).strip()
    if s == "":
        return None
    formatos = ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p", "%I%p", "%I %p")
    for fmt in formatos:
        try:
            return datetime.strptime(s, fmt).time()
        except:
            pass
    try:
        return pd.to_datetime(s, errors="coerce").time()
    except:
        return None

def generar_reporte_streamlit(wb_in, hora_inicio="09:00", hora_fin="13:30"):
    ws_data = wb_in["Data"]
    rows = [list(row) for row in ws_data.iter_rows(min_row=1, max_col=4, values_only=True)]
    if len(rows) < 2:
        st.warning("La hoja 'Data' no tiene suficientes filas.")
        return None

    headers = rows[0]
    data_rows = rows[1:]
    t_ini = pd.to_datetime(hora_inicio).time()
    t_fin = pd.to_datetime(hora_fin).time()
    parsed = []

    for idx, r in enumerate(data_rows, start=2):
        t = parse_hora_to_time(r[0])
        if t is not None and t_ini <= t <= t_fin:
            parsed.append((t, idx, r))
    parsed.sort(key=lambda x: x[0])

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Reporte"
    fecha_hoy = datetime.today().strftime("%d.%m.%Y")
    ws_out["A1"] = f"Reporte Tipo de Cambio: {fecha_hoy}"
    ws_out["A1"].font = Font(bold=True)

    for col_idx, h in enumerate(headers, start=1):
        ws_out.cell(row=2, column=col_idx, value=h)

    firstDataRowRep = 3
    lastDataRowRep = 65
    rowsPerBlock = lastDataRowRep - firstDataRowRep + 1
    need_header_copied = set()

    for i, (t, src_row_idx, row_vals) in enumerate(parsed, start=1):
        block = (i - 1) // rowsPerBlock
        rowInBlock = ((i - 1) % rowsPerBlock) + firstDataRowRep
        colStart = 1 + (block * 4)
        if block not in need_header_copied:
            for c in range(4):
                ws_out.cell(row=2, column=colStart + c, value=headers[c])
            need_header_copied.add(block)
        for c in range(4):
            ws_out.cell(row=rowInBlock, column=colStart + c, value=row_vals[c])

    cnt = len(parsed)
    if cnt > 0:
        valores_precio = np.array([p[2][1] for p in parsed], dtype=float)
        valores_monto = np.array([p[2][2] for p in parsed], dtype=float)
        vMin = round(float(np.min(valores_precio)), 4)
        vMax = round(float(np.max(valores_precio)), 4)
        vProm = round(float(np.mean(valores_precio)), 4)
        vDesv = round(float(np.std(valores_precio, ddof=0) * 1000), 4)
        vSum = int(np.nansum(valores_monto))
        filaCalc = ((cnt - 1) % rowsPerBlock) + firstDataRowRep + 1
        posColMinMax = 1 + ((cnt - 1) // rowsPerBlock) * 4
        posColMontoVol = posColMinMax + 2

        font_bold = Font(bold=True)
        ws_out.cell(row=filaCalc, column=posColMinMax, value="Mínimo").font = font_bold
        ws_out.cell(row=filaCalc, column=posColMinMax + 1, value=vMin).font = font_bold
        ws_out.cell(row=filaCalc + 1, column=posColMinMax, value="Máximo").font = font_bold
        ws_out.cell(row=filaCalc + 1, column=posColMinMax + 1, value=vMax).font = font_bold
        ws_out.cell(row=filaCalc + 2, column=posColMinMax, value="Promedio").font = font_bold
        ws_out.cell(row=filaCalc + 2, column=posColMinMax + 1, value=vProm).font = font_bold
        ws_out.cell(row=filaCalc, column=posColMontoVol, value=vSum).font = font_bold
        ws_out.cell(row=filaCalc + 1, column=posColMontoVol, value="Volatilidad").font = font_bold
        ws_out.cell(row=filaCalc + 2, column=posColMontoVol, value=vDesv).font = font_bold

    # Formato global fuera del bloque
    try:
        fuente_global = Font(name="MS Sans Serif", size=8, color="000000")
    except:
        fuente_global = Font(name="Arial", size=8, color="000000")
    alineacion_cent = Alignment(horizontal="center", vertical="center")
    for fila in range(1, ws_out.max_row + 1):
        for col in range(1, ws_out.max_column + 1):
            if cnt > 0 and filaCalc <= fila <= filaCalc + 2 and posColMinMax <= col <= posColMontoVol + 1:
                continue
            celda = ws_out.cell(row=fila, column=col)
            if celda.value not in (None, ""):
                celda.font = fuente_global
                celda.alignment = alineacion_cent

    temp_file = f"Reporte_{fecha_hoy}.xlsx"
    wb_out.save(temp_file)
    return temp_file

# ==== UI ====
st.subheader("Sube tu archivo Excel para generar el reporte")

uploaded_file = st.file_uploader(
    "Archivo Excel", type=["xlsm", "xlsx", "xls"]
)

hora_inicio = st.time_input("Hora de inicio", value=datetime.strptime("09:00", "%H:%M").time())
hora_fin = st.time_input("Hora fin", value=datetime.strptime("13:30", "%H:%M").time())

if uploaded_file is not None:
    try:
        file_ext = uploaded_file.name.split(".")[-1].lower()
        if file_ext in ["xlsm", "xlsx"]:
            wb_in = load_workbook(uploaded_file, data_only=True)
        elif file_ext == "xls":
            import xlrd
            wb_in = xlrd.open_workbook(file_contents=uploaded_file.read())
        else:
            st.error("Formato de archivo no soportado")
            wb_in = None
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}")
        wb_in = None

    if wb_in is not None and st.button("Generar Reporte"):
        ruta_reporte = generar_reporte_streamlit(wb_in, hora_inicio.strftime("%H:%M"), hora_fin.strftime("%H:%M"))
        st.success("Reporte generado correctamente ✅")
        with open(ruta_reporte, "rb") as f:
            st.download_button(
                label="Descargar Reporte",
                data=f,
                file_name=os.path.basename(ruta_reporte),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


